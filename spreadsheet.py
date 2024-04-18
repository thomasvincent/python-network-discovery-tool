import logging
import openpyxl
from typing import List

# Setup logging
logger = logging.getLogger(__name__)

class SpreadsheetManager:
    """Manages reading and writing to Excel spreadsheets."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def read_data(self) -> List[List[str]]:
        """Reads data from the active sheet and returns it as a list of lists."""
        data = []
        for row in self.sheet.iter_rows(values_only=True):
            data.append(list(row))
        return data

    def write_data(self, data: List[List[str]]) -> None:
        """Writes data to the active sheet."""
        self.sheet.delete_rows(1, self.sheet.max_row)
        for row in data:
            self.sheet.append(row)
        self.workbook.save(self.file_path)

    def add_row(self, row: List[str]) -> None:
        """Adds a single row to the active sheet."""
        self.sheet.append(row)
        self.workbook.save(self.file_path)

    def __del__(self):
        self.workbook.close()
