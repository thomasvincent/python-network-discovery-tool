"""Report service implementations.

This module provides implementations of the ReportService interface.
"""

import csv
import json
import logging
import os
from typing import List, Any

import jinja2
import openpyxl

from network_discovery.domain.device import Device
from network_discovery.application.interfaces import ReportService

# Setup logging
logger = logging.getLogger(__name__)


class ReportGenerator(ReportService):
    """Implementation of ReportService for generating reports in various formats."""

    def __init__(self, output_dir: str, template_dir: str = None) -> None:
        """Initialize a new ReportGenerator.

        Args:
            output_dir: The directory where reports will be saved.
            template_dir: The directory containing HTML templates.
        """
        self.output_dir = output_dir
        self.template_dir = template_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_report(self, devices: List[Device], format_type: str) -> Any:
        """Generate a report for a list of devices.

        Args:
            devices: The list of devices to include in the report.
            format_type: The format of the report (e.g., "html", "csv", "xlsx", "json").

        Returns:
            The path to the generated report file.

        Raises:
            ValueError: If the format type is not supported.
        """
        if format_type == "html":
            return self._generate_html_report(devices)
        elif format_type == "csv":
            return self._generate_csv_report(devices)
        elif format_type == "xlsx":
            return self._generate_excel_report(devices)
        elif format_type == "json":
            return self._generate_json_report(devices)
        else:
            raise ValueError(f"Unsupported report format: {format_type}")

    def _generate_html_report(self, devices: List[Device]) -> str:
        """Generate an HTML report.

        Args:
            devices: The list of devices to include in the report.

        Returns:
            The path to the generated HTML file.
        """
        if not self.template_dir:
            raise ValueError("Template directory is required for HTML reports")

        env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(self.template_dir),
            autoescape=jinja2.select_autoescape(["html", "xml"]),
        )
        template = env.get_template("layout.html")
        output = template.render(devices=devices)

        output_path = os.path.join(self.output_dir, "devices.html")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(output)

        logger.info("HTML report generated at %s", output_path)
        return output_path

    def _generate_csv_report(self, devices: List[Device]) -> str:
        """Generate a CSV report.

        Args:
            devices: The list of devices to include in the report.

        Returns:
            The path to the generated CSV file.
        """
        output_path = os.path.join(self.output_dir, "devices.csv")
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "Host",
                    "IP",
                    "SNMP Group",
                    "Alive",
                    "SNMP",
                    "SSH",
                    "MySQL",
                    "Username",
                    "Errors",
                ]
            )
            for device in devices:
                writer.writerow(
                    [
                        device.host,
                        device.ip,
                        device.snmp_group,
                        device.alive,
                        device.snmp,
                        device.ssh,
                        device.mysql,
                        device.uname,
                        ", ".join(device.errors),
                    ]
                )

        logger.info("CSV report generated at %s", output_path)
        return output_path

    def _generate_excel_report(self, devices: List[Device]) -> str:
        """Generate an Excel report.

        Args:
            devices: The list of devices to include in the report.

        Returns:
            The path to the generated Excel file.
        """
        output_path = os.path.join(self.output_dir, "devices.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Devices"

        # Write header
        headers = [
            "Host",
            "IP",
            "SNMP Group",
            "Alive",
            "SNMP",
            "SSH",
            "MySQL",
            "Username",
            "Errors",
        ]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # Write data
        for row, device in enumerate(devices, 2):
            sheet.cell(row=row, column=1, value=device.host)
            sheet.cell(row=row, column=2, value=device.ip)
            sheet.cell(row=row, column=3, value=device.snmp_group)
            sheet.cell(row=row, column=4, value=str(device.alive))
            sheet.cell(row=row, column=5, value=str(device.snmp))
            sheet.cell(row=row, column=6, value=str(device.ssh))
            sheet.cell(row=row, column=7, value=str(device.mysql))
            sheet.cell(row=row, column=8, value=device.uname)
            sheet.cell(row=row, column=9, value=", ".join(device.errors))

        workbook.save(output_path)
        logger.info("Excel report generated at %s", output_path)
        return output_path

    def _generate_json_report(self, devices: List[Device]) -> str:
        """Generate a JSON report.

        Args:
            devices: The list of devices to include in the report.

        Returns:
            The path to the generated JSON file.
        """
        output_path = os.path.join(self.output_dir, "devices.json")
        devices_data = [device.to_dict() for device in devices]
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(devices_data, f, indent=4)

        logger.info("JSON report generated at %s", output_path)
        return output_path
